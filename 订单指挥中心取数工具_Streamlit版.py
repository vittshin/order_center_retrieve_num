import streamlit as st
import pandas as pd
import os
import time
import re
import threading
import sys
import io
from io import StringIO
from io import BytesIO
import base64
from pdf2image import convert_from_bytes
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# 设置页面配置
st.set_page_config(
    page_title="订单指挥中心取数工具",
    page_icon="📊",
    layout="wide"
)

# 初始化session state
if 'is_processing' not in st.session_state:
    st.session_state.is_processing = False
if 'progress' not in st.session_state:
    st.session_state.progress = 0
if 'results' not in st.session_state:
    st.session_state.results = None
if 'status_text' not in st.session_state:
    st.session_state.status_text = ""

def main():
    st.title("📊 订单指挥中心取数工具")
    st.markdown("---")
    
    # 侧边栏配置
    with st.sidebar:
        st.header("⚙️ 配置设置")
        
        # 主体信息表路径
        default_data_path = r'D:\xinpingshun.1\Desktop\主体数据表.xlsx'
        data_file_path = st.text_input(
            "主体信息表路径:",
            value=default_data_path,
            help="请设置主体信息表Excel文件的路径"
        )
        
        # ChromeDriver路径
        default_driver_path = r"D:\xinpingshun.1\Desktop\chromedriver-win64\chromedriver.exe"
        driver_path = st.text_input(
            "ChromeDriver路径:",
            value=default_driver_path,
            help="请设置ChromeDriver可执行文件的路径"
        )
        
        # 输出文件路径
        default_output_path = r"D:/xinpingshun.1/Desktop/output.xlsx"
        output_path = st.text_input(
            "输出文件路径:",
            value=default_output_path,
            help="请设置输出Excel文件的路径"
        )
        
        # 图片保存路径
        default_image_path = r"D:/xinpingshun.1/Desktop/图片"
        image_path = st.text_input(
            "图片保存路径:",
            value=default_image_path,
            help="请设置图片保存的文件夹路径"
        )
        
        st.markdown("---")
        st.info("💡 提示：请确保所有路径都已正确配置，并且相关文件存在。")
    
    # 主要内容区域
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.header("📝 订单信息输入")
        order_input = st.text_area(
            "请输入订单号（用逗号、顿号、斜杠、分号、空格或制表符分隔）:",
            height=100,
            placeholder="例如：300495175125，3398207004123336，3430241009824138，3406214011263077"
        )
        
        # # 登录信息输入（如果需要）
        # st.subheader("🔐 登录信息（可选）")
        # col_user, col_pass = st.columns(2)
        # with col_user:
        #     username = st.text_input("用户名:", "")
        # with col_pass:
        #     password = st.text_input("密码:", "", type="password")
    
    with col2:
        st.header("📊 处理进度")
        progress_bar = st.progress(0)
        status_container = st.empty()
        
        # 结果预览区域
        st.subheader("📋 结果预览")
        results_placeholder = st.empty()
    
    # 开始处理按钮
    col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
    with col_btn2:
        start_button = st.button(
            "🚀 开始处理订单",
            disabled=st.session_state.is_processing,
            use_container_width=True,
            type="primary"
        )
    
    # 处理逻辑
    if start_button and order_input:
        if not os.path.exists(data_file_path):
            st.error(f"❌ 主体信息表文件不存在：{data_file_path}")
            return
        
        if not os.path.exists(os.path.dirname(output_path)):
            st.error(f"❌ 输出目录不存在：{os.path.dirname(output_path)}")
            return
        
        if not os.path.exists(image_path):
            try:
                os.makedirs(image_path)
                st.info(f"✅ 已创建图片保存目录：{image_path}")
            except Exception as e:
                st.error(f"❌ 无法创建图片保存目录：{str(e)}")
                return
        
        # 启动处理
        st.session_state.is_processing = True
        st.session_state.progress = 0
        st.session_state.status_text = "开始处理订单..."
        st.rerun()
    
    # 如果正在处理，执行处理逻辑
    if st.session_state.is_processing:
        try:
            # 读取主体信息表
            with st.spinner("正在读取主体信息表..."):
                df = pd.read_excel(data_file_path)
            
            # 设置Chrome浏览器驱动
            service = Service(executable_path=driver_path)
            options = Options()
            prefs = {
                'profile.default_content_settings.popups': 0,
                'download.default_directory': r"D:\xinpingshun.1\Desktop\发票pdf",
                "profile.default_content_setting_values.automatic_downloads": 1,
                "download.prompt_for_download": False,
                "plugins.always_open_pdf_externally": True
            }
            options.add_experimental_option('prefs', prefs)
            options.add_argument("--headless")
            options.add_argument("--disable-gpu")
            options.add_experimental_option('excludeSwitches', ['enable-automation'])
            options.add_argument("--no-sandbox");
            options.add_argument("--disable-dev-shm-usage");
            options.add_argument("--window-size=1920,1080");	#建议设置窗口大小

            options.add_argument("--unsafely-treat-insecure-origin-as-secure=http://order.control.jd.com/index?url=hello&code=_pQob5rsXq2ipCq1lke-FxtquCM_Ui3VHCgB4gmZDTc&state=3lMx0Ip5UHU5ycJ8ZngTAVaD-FFI1UjOIrian6ZhgjQ#/query/om")
            
            driver = webdriver.Chrome(service=service, options=options)
            driver.get("http://order.control.jd.com/index?url=hello&code=ASwBCf5Pl5bxj-ee6MczW9iZMF80ibhMIC-KSwhWXeQ&state=W9HJEjPaorg6QXodeIZenHr6iFi9tJjaY1TXtESgC5c#/query/om")
            time.sleep(1)
            
            # # 如果提供了用户名和密码，执行登录 /html/body/div[1]/div[1]/div[1]/div[1]/form/div[2]/label/div/input
            # if username and password:
            #     try:
            #         with st.spinner("正在登录..."):
            #             time.sleep(1)
            #             driver.find_element(By.XPATH, "/html/body/div[1]/div[1]/div[1]/div[1]/form/div[2]/label/div/input").send_keys(username)
            #             time.sleep(1)
            #             driver.find_element(By.XPATH, "/html/body/div[1]/div[1]/div[1]/div[1]/form/div[3]/label/div/input").send_keys(password)
            #             time.sleep(1)
            #             driver.find_element(By.XPATH, "/html/body/div[1]/div[1]/div[1]/div[5]/div[2]/input").click()
            #             time.sleep(30)
            #     except Exception as e:
            #         st.warning(f"⚠️ 登录失败：{str(e)}")
            
            # 订单号拆分
            order_split = re.split(r",|，|/|;|\s|\t", order_input)
            
            # 初始化数据列表
            orders_no = []
            child_orders_no = []
            reply_content = []
            fapiao_url = []
            is_parent = []
            orders_type = []
            image_data_list = []
            
            def switch_to_frame():
                driver.switch_to.default_content()
                try:
                    driver.switch_to.frame(0)
                except:
                    driver.switch_to.frame(1)
            
            pic_row = 1
            excel_path = output_path
            wb = openpyxl.Workbook()
            wb.save(excel_path)
            
            # 写入图片函数
            def write_image(excel_path, img_path_1, img_path_2):
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
                img_1 = Image(img_path_1)
                img_2 = Image(img_path_2)
                img_1.width, img_1.height = (80, 80)
                img_2.width, img_2.height = (80, 80)
                ws.column_dimensions[img_column_1].width = 20
                ws.column_dimensions[img_column_2].width = 20
                ws.row_dimensions[pic_row].height = 85
                ws.add_image(img_1, anchor=img_column_1 + str(pic_row))
                ws.add_image(img_2, anchor=img_column_2 + str(pic_row))
                wb.save(excel_path)
                wb.close()
            
            # 开始处理订单 - 这里是第101-532行的核心逻辑
            total_orders = len(order_split)
            progress_step = 1.0 / total_orders if total_orders > 0 else 1.0
            
            for order_item in order_split:
                if not order_item.strip():
                    continue
                
                order_item = order_item.strip()
                
                # 更新状态
                st.session_state.status_text = f"正在处理订单：{order_item}"
                status_container.info(st.session_state.status_text)
                progress_bar.progress(st.session_state.progress)
                
                time.sleep(10)
                switch_to_frame()
                
                # 搜索框元素定位
                search_input = driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/span[2]/input[1]")
                time.sleep(1)
                search_input.clear()
                time.sleep(1)
                search_input.send_keys(order_item)
                
                # 定位按钮元素
                time.sleep(1)
                button_element = driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/span[2]/input[2]")
                
                # 点击按钮
                time.sleep(1)
                button_element.click()
                
                # 判断是否为父单
                time.sleep(1)
                try:
                    order_type = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/table[1]/tbody/tr[9]/td[2]").text
                except:
                    order_type = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/table[1]/tbody/tr[9]/td[2]").text
                time.sleep(1)
                try:
                    institude_name = driver.find_element(By.XPATH, "/html/body/div[1]/div[1]/span[1]").text
                except:
                    institude_name = driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/span[1]").text
                
                if '子单' in order_type:
                    pic_row += 1
                    orders_no.append(order_item)
                    child_orders_no.append("")
                    is_parent.append("子单")
                    
                    # 判断是否为POP机构
                    if 'POP机构' in institude_name:
                        orders_type.append("POP订单")
                        fapiao_url.append("")
                        
                        switch_to_frame()
                        handles = driver.window_handles
                        driver.switch_to.window(handles[-1])
                        time.sleep(1)
                        shopper_id = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[23]/table/tbody/tr[2]/td[2]").text
                        js = "window.open('{}','——blank');"
                        driver.execute_script(js.format('http://yy.jd.com/vender/venderList'))
                        time.sleep(10)
                        
                        handles = driver.window_handles
                        driver.switch_to.window(handles[-1])
                        time.sleep(1)
                        driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[1]/div[2]/div/span[2]").click()
                        time.sleep(1)
                        driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[2]/div[2]/div[3]/div").click()
                        time.sleep(2)
                        driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div/div/div[3]/form/div[1]/div/div/input").send_keys(shopper_id)
                        time.sleep(1)
                        driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div/div/div[4]/div[1]/div/div[1]/div/div/div/div[3]").click()
                        time.sleep(1)
                        driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div/div/div[3]/form/div[9]/div/button[1]/span").click()
                        time.sleep(1)
                        driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div/div/div[4]/div[2]/div/div[2]/div[5]/div[2]/table/tbody/tr/td[11]/div/div/div/div[1]").click()
                        time.sleep(1)
                        driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[1]/div/div[1]/div/div/div/div[2]").click()
                        time.sleep(1)
                        shop_info = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/div[1]/div[2]/form/div[5]/div/div").text
                        
                        if "POP自然人店" in shop_info:
                            time.sleep(1)
                            driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[1]/div/div[1]/div/div/div/div[5]").click()
                            time.sleep(1)
                            name = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[3]/div[2]/div[3]/div[2]/div/div").text
                            time.sleep(1)
                            id_card = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[3]/div[2]/div[4]/div[2]/div/span").text
                            reply_content.append(f"订单号：{order_item}\n姓名：{name}\n籍贯： \n身份证号：{id_card}")
                            
                            # 提取并处理图片
                            headers = {
                                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
                            }
                            time.sleep(1)
                            img_url_1 = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[3]/div[2]/div[2]/div[2]/div/div[1]/div[1]/img").get_attribute("src")
                            img_url_2 = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[3]/div[2]/div[2]/div[2]/div/div[1]/div[2]/img").get_attribute("src")
                            
                            response_1 = requests.get(img_url_1, headers=headers)
                            response_2 = requests.get(img_url_2, headers=headers)
                            image_data_1 = response_1.content
                            image_data_2 = response_2.content
                            with open(f"{image_path}/{order_item}_1.jpg", "wb") as f:
                                f.write(image_data_1)
                            with open(f"{image_path}/{order_item}_2.jpg", "wb") as f:
                                f.write(image_data_2)
                            
                            img_column_1 = 'G'
                            img_column_2 = 'H'
                            img_path_1 = f"{image_path}/{order_item}_1.jpg"
                            img_path_2 = f"{image_path}/{order_item}_2.jpg"
                            
                            write_image(excel_path, img_path_1, img_path_2)
                            
                            handles = driver.window_handles
                            driver.switch_to.window(handles[-1])
                            driver.close()
                            driver.switch_to.window(handles[0])
                        
                        elif "POP-SOP" in shop_info:
                            time.sleep(1)
                            driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[1]/div/div[1]/div/div/div/div[5]").click()
                            time.sleep(1)
                            comp_name = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[2]/div[2]/div[2]/div[2]/div/div").text
                            time.sleep(1)
                            tax_no = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[2]/div[2]/div[1]/div[2]/div/span").text
                            time.sleep(1)
                            reg_place = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[2]/div[2]/div[16]/div[2]/div/span").text
                            if "北京" in reg_place:
                                reply_content.append(f"订单号：{order_item}\n公司名称：{comp_name}\n税号：{tax_no}\n注册地：{reg_place}\n北京第三方")
                            else:
                                reply_content.append(f"订单号：{order_item}\n公司名称：{comp_name}\n税号：{tax_no}\n注册地：{reg_place}\n外省第三方")
                            
                            handles = driver.window_handles
                            driver.switch_to.window(handles[-1])
                            driver.close()
                            driver.switch_to.window(handles[0])
                        
                        else:
                            time.sleep(1)
                            driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[1]/div/div[1]/div/div/div/div[5]").click()
                            time.sleep(1)
                            comp_name = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[2]/div[2]/div[2]/div[2]/div/div").text
                            time.sleep(1)
                            reg_place = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[2]/div[2]/div[4]/div[2]/div/span").text
                            reply_content.append(f"订单号：{order_item}\n海外主体名称：{comp_name}\n公司注册地：{reg_place}\n全球购")
                            
                            handles = driver.window_handles
                            driver.switch_to.window(handles[-1])
                            driver.close()
                            driver.switch_to.window(handles[0])
                    
                    # 判断是否为自营-香港
                    elif '香港international' in institude_name:
                        orders_type.append("全球购")
                        fapiao_url.append("")
                        reply_content.append(f"订单号：{order_item}\n全球购\nJD.com international limited\n香港")

                    
                    # 判断是否为自营-非香港
                    else:
                        time.sleep(1)
                        try:
                            fapiao_issue_name = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[18]/table/tbody/tr[3]/td[6]").text
                        except:
                            fapiao_issue_name = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[18]/table/tbody/tr[3]/td[6]").text
                        regis_posit = df.loc[df.loc[df['贸易公司简称'] == fapiao_issue_name].index[0], '注册地区']
                        comp_name = df.loc[df.loc[df['贸易公司简称'] == fapiao_issue_name].index[0], '公司名称']
                        tax_no = df.loc[df.loc[df['贸易公司简称'] == fapiao_issue_name].index[0], '统一社会信用代码']
                        
                        if '北京市' in regis_posit:
                            orders_type.append("北京自营")
                            reply_content.append(f"订单号：{order_item}\n公司名称：{comp_name}\n税号：{tax_no}\n北京自营")
                            time.sleep(1)
                            switch_to_frame()
                            try:
                                url = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[18]/table/tbody/tr[8]/td[2]/a[last()]").get_attribute("href")
                            except:
                                url = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[18]/table/tbody/tr[8]/td[2]/a[last()]").get_attribute("href")
                            fapiao_url.append(url)
                        else:
                            orders_type.append("外省自营")
                            reply_content.append(f"订单号：{order_item}\n公司名称：{comp_name}\n税号：{tax_no}\n外省自营")
                            fapiao_url.append("")
                
                else:
                    switch_to_frame()
                    time.sleep(1)
                    driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/table[1]/tbody/tr[9]/td[2]/a").click()
                    parent_element = driver.find_element(By.XPATH, "/html/body/div[4]/div[2]")
                    time.sleep(3)
                    child_elements = parent_element.find_elements(By.XPATH, "./a")
                    child_count = len(child_elements) - 1
                    
                    for child in range(child_count):
                        switch_to_frame()
                        time.sleep(1)
                        driver.find_element(By.XPATH, f"/html/body/div[4]/div[2]/a[{child+2}]").click()
                        time.sleep(1)
                        child_orders_no.append(driver.find_element(By.XPATH, f"/html/body/div[4]/div[2]/a[{child+2}]").text)
                        orders_no.append(order_item)
                        pic_row += 1
                        is_parent.append("父单")
                        
                        handles = driver.window_handles
                        driver.switch_to.window(handles[-1])
                        
                        time.sleep(1)
                        try:
                            order_type = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/table[1]/tbody/tr[9]/td[2]").text
                        except:
                            order_type = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/table[1]/tbody/tr[9]/td[2]").text
                        time.sleep(1)
                        try:
                            institude_name = driver.find_element(By.XPATH, "/html/body/div[1]/div[1]/span[1]").text
                        except:
                            institude_name = driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/span[1]").text
                        
                        if 'POP机构' in institude_name:
                            orders_type.append("POP订单")
                            fapiao_url.append("")
                            
                            switch_to_frame()
                            handles = driver.window_handles
                            driver.switch_to.window(handles[-1])
                            time.sleep(1)
                            shopper_id = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[23]/table/tbody/tr[2]/td[2]").text
                            js = "window.open('{}','——blank');"
                            driver.execute_script(js.format('http://yy.jd.com/vender/venderList'))
                            time.sleep(10)
                            
                            handles = driver.window_handles
                            driver.switch_to.window(handles[-1])
                            time.sleep(1)
                            driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[1]/div[2]/div/span[2]").click()
                            time.sleep(1)
                            driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[2]/div[2]/div[3]/div").click()
                            time.sleep(2)
                            driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div/div/div[3]/form/div[1]/div/div/input").send_keys(shopper_id)
                            time.sleep(1)
                            driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div/div/div[4]/div[1]/div/div[1]/div/div/div/div[3]").click()
                            time.sleep(1)
                            driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div/div/div[3]/form/div[9]/div/button[1]/span").click()
                            time.sleep(1)
                            driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div/div/div[4]/div[2]/div/div[2]/div[5]/div[2]/table/tbody/tr/td[11]/div/div/div/div[1]").click()
                            time.sleep(1)
                            driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[1]/div/div[1]/div/div/div/div[2]").click()
                            time.sleep(1)
                            shop_info = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/div[1]/div[2]/form/div[5]/div/div").text
                            
                            if "POP自然人店" in shop_info:
                                time.sleep(1)
                                driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[1]/div/div[1]/div/div/div/div[5]").click()
                                time.sleep(1)
                                name = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[3]/div[2]/div[3]/div[2]/div/div").text
                                time.sleep(1)
                                id_card = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[3]/div[2]/div[4]/div[2]/div/span").text
                                reply_content.append(f"订单号：{order_item}\n姓名：{name}\n籍贯： \n身份证号：{id_card}")
                                
                                headers = {
                                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
                                }
                                time.sleep(1)
                                img_url_1 = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[3]/div[2]/div[2]/div[2]/div/div[1]/div[1]/img").get_attribute("src")
                                img_url_2 = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[3]/div[2]/div[2]/div[2]/div/div[1]/div[2]/img").get_attribute("src")
                                
                                response_1 = requests.get(img_url_1, headers=headers)
                                response_2 = requests.get(img_url_2, headers=headers)
                                image_data_1 = response_1.content
                                image_data_2 = response_2.content
                                with open(f"{image_path}/{order_item}_1.jpg", "wb") as f:
                                    f.write(image_data_1)
                                with open(f"{image_path}/{order_item}_2.jpg", "wb") as f:
                                    f.write(image_data_2)
                                
                                img_column_1 = 'G'
                                img_column_2 = 'H'
                                img_path_1 = f"{image_path}/{order_item}_1.jpg"
                                img_path_2 = f"{image_path}/{order_item}_2.jpg"
                                
                                write_image(excel_path, img_path_1, img_path_2)
                                
                                handles = driver.window_handles
                                driver.switch_to.window(handles[-1])
                                driver.close()
                                handles = driver.window_handles
                                driver.switch_to.window(handles[-1])
                                driver.close()
                                driver.switch_to.window(handles[0])
                            
                            elif "POP-SOP" in shop_info:
                                time.sleep(1)
                                driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[1]/div/div[1]/div/div/div/div[5]").click()
                                time.sleep(1)
                                comp_name = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[2]/div[2]/div[2]/div[2]/div/div").text
                                time.sleep(1)
                                tax_no = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[2]/div[2]/div[1]/div[2]/div/span").text
                                time.sleep(1)
                                reg_place = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[2]/div[2]/div[16]/div[2]/div/span").text
                                if "北京" in reg_place:
                                    reply_content.append(f"订单号：{order_item}\n公司名称：{comp_name}\n税号：{tax_no}\n注册地：{reg_place}\n北京第三方")
                                else:
                                    reply_content.append(f"订单号：{order_item}\n公司名称：{comp_name}\n税号：{tax_no}\n注册地：{reg_place}\n外省第三方")
                                
                                handles = driver.window_handles
                                driver.switch_to.window(handles[-1])
                                driver.close()
                                handles = driver.window_handles
                                driver.switch_to.window(handles[-1])
                                driver.close()
                                driver.switch_to.window(handles[0])
                            
                            else:
                                time.sleep(1)
                                driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[1]/div/div[1]/div/div/div/div[5]").click()
                                time.sleep(1)
                                comp_name = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[2]/div[2]/div[2]/div[2]/div/div").text
                                time.sleep(1)
                                reg_place = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/micro-app/micro-app-body/div[1]/div/div[5]/div/div/div/section/div[2]/div[2]/div[2]/form/div[2]/div[2]/div[4]/div[2]/div/span").text
                                reply_content.append(f"订单号：{order_item}\n海外主体名称：{comp_name}\n公司注册地：{reg_place}\n全球购")
                                
                                handles = driver.window_handles
                                driver.switch_to.window(handles[-1])
                                driver.close()
                                handles = driver.window_handles
                                driver.switch_to.window(handles[-1])
                                driver.close()
                                driver.switch_to.window(handles[0])
                        
                        # 判断是否为自营-香港
                        elif '香港international' in institude_name:
                            orders_type.append("全球购")
                            fapiao_url.append("")
                            reply_content.append(f"订单号：{order_item}\n全球购\nJD.com international limited\n香港")
                            
                            driver.close()
                            driver.switch_to.window(handles[0])
                        
                        # 判断是否为自营-非香港
                        else:
                            time.sleep(1)
                            try:
                                fapiao_issue_name = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[18]/table/tbody/tr[3]/td[6]").text
                            except:
                                fapiao_issue_name = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[18]/table/tbody/tr[3]/td[6]").text
                            regis_posit = df.loc[df.loc[df['贸易公司简称'] == fapiao_issue_name].index[0], '注册地区']
                            comp_name = df.loc[df.loc[df['贸易公司简称'] == fapiao_issue_name].index[0], '公司名称']
                            tax_no = df.loc[df.loc[df['贸易公司简称'] == fapiao_issue_name].index[0], '统一社会信用代码']
                            
                            if '北京市' in regis_posit:
                                orders_type.append("北京自营")
                                reply_content.append(f"订单号：{order_item}\n公司名称：{comp_name}\n税号：{tax_no}\n北京自营")
                                time.sleep(1)
                                switch_to_frame()
                                try:
                                    url = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[18]/table/tbody/tr[8]/td[2]/a[last()]").get_attribute("href")
                                except:
                                    url = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[18]/table/tbody/tr[8]/td[2]/a[last()]").get_attribute("href")
                                fapiao_url.append(url)
                                
                                driver.close()
                                driver.switch_to.window(handles[0])
                            else:
                                orders_type.append("外省自营")
                                reply_content.append(f"订单号：{order_item}\n公司名称：{comp_name}\n税号：{tax_no}\n外省自营")
                                fapiao_url.append("")
                                
                                driver.close()
                                driver.switch_to.window(handles[0])
                
                # 更新进度
                st.session_state.progress += progress_step
                progress_bar.progress(min(st.session_state.progress, 1.0))
                
                # 显示当前结果预览
                if len(orders_no) > 0:
                    preview_df = pd.DataFrame({
                        '订单号': orders_no,
                        '子订单号': child_orders_no,
                        '订单类型': orders_type,
                        '是否为父单': is_parent,
                        '回复内容': reply_content,
                        '发票号链接': fapiao_url
                    })
                    results_placeholder.dataframe(preview_df, use_container_width=True)
            
            # 关闭浏览器
            driver.quit()
            
            # 创建最终数据框
            df_output = {
                '订单号': orders_no,
                '子订单号': child_orders_no,
                '订单类型': orders_type,
                '是否为父单': is_parent,
                '回复内容': reply_content,
                '发票号链接': fapiao_url
            }
            df_output = pd.DataFrame(df_output)
            
            # 保存到Excel
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            ws.append(df_output.columns.to_list())
            for row in range(len(df_output)):
                ws.append(df_output.loc[row].to_list())
            wb.save(excel_path)
            
            # 更新状态
            st.session_state.status_text = "处理完成！"
            status_container.success(st.session_state.status_text)
            progress_bar.progress(1.0)
            st.session_state.results = df_output
            st.session_state.is_processing = False
            
            # 显示成功消息
            st.success(f"✅ 处理完成！共处理了 {len(order_split)} 个订单，结果已保存到：{excel_path}")
            
            # 提供下载按钮
            with open(excel_path, "rb") as file:
                st.download_button(
                    label="📥 下载结果文件",
                    data=file,
                    file_name=os.path.basename(excel_path),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
        except Exception as e:
            st.session_state.is_processing = False
            st.error(f"❌ 处理过程中出现错误：{str(e)}")
            if 'driver' in locals():
                driver.quit()
    
    # 显示历史结果
    if st.session_state.results is not None and not st.session_state.is_processing:
        st.markdown("---")
        st.header("📊 处理结果")
        st.dataframe(st.session_state.results, use_container_width=True)
        
        # 统计信息
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("总订单数", len(st.session_state.results))
        with col2:
            st.metric("父单数", len(st.session_state.results[st.session_state.results['是否为父单'] == '父单']))
        with col3:
            st.metric("子单数", len(st.session_state.results[st.session_state.results['是否为父单'] == '子单']))
        with col4:
            order_types = st.session_state.results['订单类型'].value_counts()
            st.metric("订单类型数", len(order_types))

if __name__ == "__main__":
    main()